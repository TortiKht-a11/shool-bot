"""Excel exporter for applications."""

from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from .texts import STATUS_RU


HEADERS = [
    "ID",
    "Дата создания",
    "Статус",
    "ФИО ребёнка",
    "Дата рождения",
    "Пол",
    "Адрес проживания",
    "Адрес регистрации",
    "Детский сад",
    "ФИО родителя",
    "Родство",
    "Телефон",
    "Email",
    "Место работы",
    "Комментарий админа",
]


def export_applications_xlsx(rows: list[dict], out_dir: Path) -> Path:
    out_dir.mkdir(parents=True, exist_ok=True)
    file_name = f"export_{datetime.now().strftime('%Y-%m-%d_%H-%M')}.xlsx"
    path = out_dir / file_name

    wb = Workbook()
    ws = wb.active
    ws.title = "Заявки"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="2F5597")
    center = Alignment(vertical="center", wrap_text=True)

    ws.append(HEADERS)
    for col, _ in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for r in rows:
        ws.append(
            [
                r.get("id"),
                r.get("created_at"),
                STATUS_RU.get(r.get("status", ""), r.get("status")),
                r.get("child_full_name"),
                r.get("child_birth_date"),
                r.get("child_gender"),
                r.get("child_address"),
                r.get("child_registration_address"),
                r.get("kindergarten") or "",
                r.get("parent_full_name"),
                r.get("parent_relation"),
                r.get("parent_phone"),
                r.get("parent_email") or "",
                r.get("parent_work") or "",
                r.get("admin_comment") or "",
            ]
        )

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(HEADERS)):
        for cell in row:
            cell.alignment = center

    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            val = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(val))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 60)

    wb.save(path)
    return path
